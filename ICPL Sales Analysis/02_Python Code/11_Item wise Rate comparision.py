import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime

# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# ✅ Remove Quantity = 0
data = data[data['Quantity'] > 0]

# Group by Item Name and calculate required metrics
grouped = data.groupby('Item Name').agg(
    Max_Rate=('Rate', 'max'),
    Min_Rate=('Rate', 'min')
).reset_index()

# Add Max - Min column
grouped['Max - Min'] = grouped['Max_Rate'] - grouped['Min_Rate']

# ✅ Filter where difference is greater than 1
grouped = grouped[grouped['Max - Min'] > 1]

grouped = grouped.sort_values(by='Max - Min', ascending=False)

# ✅ Rename columns as per your requirement
grouped.rename(columns={
    'Max_Price': 'Max Rate',
    'Min_Price': 'Min Rate'
}, inplace=True)

# Add Sr. No. column
grouped.insert(0, 'SR. No.', range(1, len(grouped) + 1))

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/11_Item wise Rate comparision.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    grouped.to_excel(writer, sheet_name=' Rate comparision', index=False)

print("✅ Data successfully saved to new file:", output_path)