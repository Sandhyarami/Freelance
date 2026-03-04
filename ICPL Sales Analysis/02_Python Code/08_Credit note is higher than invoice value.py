import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime

# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Credit Not Register.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Credit Note Register']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Credit Note Register', engine='openpyxl')

voucher_ref_list = (
    data['Voucher Ref. No.']
    .dropna()
    .astype(str)
    .str.strip()
    .unique()
)

print(voucher_ref_list)

# ✅ Remove blank Voucher Ref. No.
data = data[data['Voucher Ref. No.'].notna() & (data['Voucher Ref. No.'].astype(str).str.strip() != '')]

# 🔹 Step 3: Load second file
file_path2 = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx"
data2 = pd.read_excel(file_path2, sheet_name='Working', engine='openpyxl')  # Replace with actual sheet if different

# ✅ Step 4: Filter rows where 'Voucher No.' in second file matches any in voucher_ref_list
data2['Voucher No.'] = data2['Voucher No.'].astype(str).str.strip()
matched_data = data2[data2['Voucher No.'].isin(voucher_ref_list)]

# 🔹 Step 5: Output result
print(matched_data)

# ✅ Group by Date, Customer, Voucher No. and sum Values
summary = (
    matched_data.groupby(['Date', 'Customer', 'Voucher No.'], dropna=False)['Value']
    .sum()
    .reset_index()
    .sort_values(by='Date')
)

print(summary)

# 🔁 Step 5: Merge into full Credit Register
merged_result = pd.merge(
    data,        # full original data, not just filtered
    summary,            # summary from matched sales data
    how='left',
    left_on='Voucher Ref. No.',
    right_on='Voucher No.'
)
merged_result.drop(columns=['Voucher No._y'], inplace=True, errors='ignore')

# 🏷 Rename columns
merged_result.rename(columns={
    'Date_y': 'Sales Date',
    'Customer_y': 'Customer Name',
    'Voucher No._x' : 'Voucher No.',
    'Value_x': 'Sales Return Value',
    'Value_y' : 'Sales Value',
    'Customer_x' : 'Customer',
    'Date_x' : 'Date'
    # Add any other columns here...
}, inplace=True)

merged_result.drop(columns=['Customer'], inplace=True, errors='ignore')
merged_result.drop(columns=['Sales Date'], inplace=True, errors='ignore')

# ➕ Add Value Difference column
merged_result['Value Difference'] = merged_result['Sales Value'] - merged_result['Sales Return Value']

print(merged_result)

# 🔍 Filter rows where Value Difference is less than 0
filtered_result = merged_result[merged_result['Value Difference'] < 0]

# ✅ Optional: Save result
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/08_Credit note is higher than invoice value.xlsx"
filtered_result.to_excel(output_path, sheet_name='Credit Note',index=False)