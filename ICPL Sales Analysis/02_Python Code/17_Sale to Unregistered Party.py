import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime

# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# ✅ List of customers to remove
remove_customers = [
    "CHETAK LOS ANGELES LLC",
    "Chetak Chicago Llc",
    "Chetak Melbourne Pvt Ltd",
    "Chetak New York Llc",
    "Chetak Orlando Llc",
    "Chetak Sanfrancisco Llc",
    "Deep Canada Inc.",
    "Deep Canada Inc. SERVICES",
    "Deep Foods (z)",
    "Deep Foods Inc.",
    "Deepkiran Foods Pvt Ltd"
]

# ✅ Remove rows where 'Customer' matches any name in the list
data = data[~data['Customer'].isin(remove_customers)]

print("✅ Filtered data (after removing specific customers):")
print(data)

# Filter records where GST column is blank or NaN
blank_gst_df = data[data['GSTIN/UIN'].isna() | (data['GSTIN/UIN'].astype(str).str.strip() == '')]

# Show the result
print(blank_gst_df)

# ✅ Group by Customer and calculate sum of 'Value'
grouped = blank_gst_df.groupby('Customer', as_index=False)['Value'].sum()

# ✅ Add Sr No.
grouped.insert(0, 'Sr No.', range(1, len(grouped) + 1))

# Show the grouped summary
print(grouped)

# Save to single Excel sheet
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/17_Sale to Unregistered Party.xlsx"
grouped.to_excel(output_path, index=False, sheet_name='Unregistered Party')

print(f"✅ Only matched item records saved to: {output_path}")