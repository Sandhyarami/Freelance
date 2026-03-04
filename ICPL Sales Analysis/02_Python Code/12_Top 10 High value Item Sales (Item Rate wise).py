import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
import re
# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# Grouping by 'Date', 'Item Details', and 'Voucher No.'
grouped_data = data.groupby(['Customer','Item Name']).agg(
    Rate_sum=('Rate', 'sum'),
    Value_sum=('Value', 'sum')
).reset_index()

# Calculate total amount
total_amount = grouped_data['Value_sum'].sum()

# Sort by Value_sum in ascending order
grouped_data = grouped_data.sort_values(by='Value_sum', ascending=False)

# Add Sr. No. column
grouped_data.insert(0, 'SR. No.', range(1, len(grouped_data) + 1))

# Keep only rows where Sr. No. is between 1 and 10
grouped_data = grouped_data[grouped_data['SR. No.'] <= 10]

# Append total row
total_row = pd.DataFrame({
    'SR. No.': [np.nan],
    'Customer': ['Total'],
    'Value_sum': [total_amount],  # Total Amount remains as is
})

grouped_data = pd.concat([grouped_data, total_row], ignore_index=True)

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/12_Top 10 High value Item Sales (Item Rate wise).xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    grouped_data.to_excel(writer, sheet_name='Top 10 Customer', index=False)

print("✅ Data successfully saved to new file:", output_path)
