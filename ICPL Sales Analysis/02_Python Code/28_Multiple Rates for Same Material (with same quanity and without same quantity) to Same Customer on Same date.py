import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
from collections import OrderedDict

# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
df = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

# === Clean column names ===
df.columns = df.columns.str.strip()

# === Required columns ===
required_columns = ['Date', 'Customer', 'Item Name', 'Rate', 'Quantity', 'Voucher Type']
missing = [col for col in required_columns if col not in df.columns]
if missing:
    raise ValueError(f"Missing columns: {missing}")

# === Ensure Date column is in datetime format ===
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

# === Group the data ===
grouped = df.groupby(['Date', 'Customer', 'Item Name', 'Voucher Type'])

report_records = []

# === Analyze groups with multiple rates ===
for (date, Customer, item, voucher), group in grouped:
    unique_rates = group['Rate'].dropna().unique()

    if len(unique_rates) > 1:
        # Group and count Rates in original order
        rate_counts = OrderedDict()
        for rate in group['Rate']:
            rate_counts[rate] = rate_counts.get(rate, 0) + 1
        rate_count_str = ' | '.join([f"{rate}({count})" for rate, count in rate_counts.items()])

        # Group and count Quantities in original order
        qty_counts = OrderedDict()
        for qty in group['Quantity']:
            qty_counts[qty] = qty_counts.get(qty, 0) + 1
        qty_count_str = ' | '.join([f"{qty}({count})" for qty, count in qty_counts.items()])

        total_quantity = group['Quantity'].sum()
        same_quantity_flag = len(qty_counts) == 1

        report_records.append({
            'Date': date,
            'Vendor': Customer,
            'Item Name': item,
            'Voucher Type': voucher,
            'Count of Rate': len(rate_counts),
            'List of Rate with Count': rate_count_str,
            'List of Quantity with Count': qty_count_str,
            'Total Quantity': total_quantity,
            'Difference': max(unique_rates) - min(unique_rates),
            'Same Quantity': same_quantity_flag
        })

# === Create DataFrame ===
df_report = pd.DataFrame(report_records)

# Add Sr.No.
df_report.index += 1
df_report.reset_index(inplace=True)
df_report.rename(columns={'index': 'Sr.No.'}, inplace=True)

# Sort by Difference descending
df_report = df_report.sort_values(by='Difference', ascending=False)

# Split into two sheets based on Same Quantity
df_same_qty = df_report[df_report['Same Quantity'] == True].drop(columns=['Same Quantity'])
df_diff_qty = df_report[df_report['Same Quantity'] == False].drop(columns=['Same Quantity'])

# === Output file path ===
output_file = "D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/28_Multiple Rates for Same Material (with same quanity and without same quantity) to Same Customer on Same date.xlsx"

# Write to Excel with two sheets
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_same_qty.to_excel(writer, sheet_name='Same Quantity', index=False)
    df_diff_qty.to_excel(writer, sheet_name='Different Quantity', index=False)

print(f"✅ Report saved at: {output_file}")
