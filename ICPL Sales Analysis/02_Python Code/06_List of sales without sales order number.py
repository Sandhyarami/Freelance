import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
import re
# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# ✅ Filter rows where "Order No." is blank (NaN or empty string)
blank_order_data = data[data['Order No.'].isna() | (data['Order No.'].astype(str).str.strip() == '')]

# 🔢 Add SR No. starting from 1
blank_order_data.insert(0, 'SR No.', range(1, len(blank_order_data) + 1))

# ✅ List of required columns
required_columns = [
    'SR No.', 'Date', 'Customer', 'Item Name', 'Voucher Type', 'Voucher No.',
    'Sales Type', 'GSTIN/UIN', 'PAN No.', 'Order No.', 'Quantity',
    'Rate', 'Value', 'Order Date'
]

# ✅ Select only required columns
blank_order_data = blank_order_data[required_columns]

print("✅ Blank Order No. records:")
print(blank_order_data)

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/06_List of sales without sales order number.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    blank_order_data.to_excel(writer, sheet_name='Without Order No.', index=False)

print("✅ Data successfully saved to new file:", output_path)