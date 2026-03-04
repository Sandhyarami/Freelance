import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
import re
# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# Replace with actual column names
category_col = 'Product Category'   # Product category column  

data[category_col] = data[category_col].fillna('Blank') 

# Group by Product Category and calculate count and sum
summary = data.groupby(category_col).agg(
    NO_OF_RECS=('Product Category', 'count'),
    VALUES=('Value', 'sum')
).reset_index()

summary = summary.sort_values(by='VALUES', ascending=False).reset_index(drop=True)

# Add Sr. No.
summary.insert(0, 'SR No.', range(1, len(summary) + 1))

# Format VALUE_SUM with commas and 2 decimal places
summary['VALUES'] = summary['VALUES'].apply(lambda x: f"{x:,.2f}")

# ✅ Output
print(summary)

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/15_Sales Mix (Product category wise).xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    summary.to_excel(writer, sheet_name='Sales Mix', index=False)

print("✅ Data successfully saved to new file:", output_path)
