import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime

# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# ✅ Remove Quantity = 0
data = data[data['Quantity'] > 0]

# Ensure 'Rate' is numeric
data['Rate'] = pd.to_numeric(data['Rate'], errors='coerce')

# Group by Date, Item Name, Voucher No
grouped = data.groupby(['Date', 'Item Name', 'Voucher No.'])

# Create a new DataFrame with Rate count and unique list
summary = grouped['Rate'].agg([
    ('Unique Rate Count', lambda x: x.nunique()),
    ('Unique Rate List', lambda x: ' | '.join(map(str, sorted(x.dropna().unique()))))
]).reset_index()

summary = summary[summary['Unique Rate Count'] >= 2]

# ✅ Rename columns as per your requirement
summary.rename(columns={
    'Unique Rate Count': 'Count of Rate',
    'Unique Rate List': 'List of Rate'
}, inplace=True)

# 🔢 Add SR No. starting from 1
summary.insert(0, 'SR No.', range(1, len(summary) + 1))

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/13_Sale Rate analysis within Same Invoice.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    summary.to_excel(writer, sheet_name='Same Invoice', index=False)

print("✅ Data successfully saved to new file:", output_path)