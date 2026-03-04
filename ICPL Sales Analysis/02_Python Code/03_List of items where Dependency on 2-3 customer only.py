import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
import re
# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# ✅ List of customers to remove
remove_customers = [
    "CHETAK LOS ANGELES LLC",
    "Chetak Chicago Llc",
    "Chetak Melbourne Pvt Ltd",
    "Chetak New York Llc",
    "Chetak Orlando Llc",
    "Chetak Sanfrancisco Llc",
    "Deep Canada Inc.",
    "Deep Canada Inc. SERVICES",
    "Deep Foods (z)",
    "Deep Foods Inc.",
    "Deepkiran Foods Pvt Ltd"
]

# ✅ Remove rows where 'Customer' matches any name in the list
data = data[~data['Customer'].isin(remove_customers)]

print("✅ Filtered data (after removing specific customers):")
print(data)

# Grouping by 'Date', 'Item Details', and 'Voucher No.'
grouped_data = data.groupby(['Item Name']).agg(
    Customer_count=('Customer', pd.Series.nunique),  # Unique customer count
    Customer_list=('Customer', lambda x: ' / '.join(pd.Series(x.unique()).astype(str))) ,
    Amount_sum=('Value', 'sum')
).reset_index()

# Filter rows where Amount_count is either 2 or 3
grouped_data = grouped_data[(grouped_data['Customer_count'].isin([2, 3]))]

# ✅ Rename columns as per your requirement
grouped_data.rename(columns={
    'Customer_count': 'Customer Count',
    'Customer_list': 'List of Customers',
    'Amount_sum' : 'Total Sales Value'
}, inplace=True)

# ✅ Filter where Total Sales Value > 100000
grouped_data = grouped_data[grouped_data['Total Sales Value'] > 100000]

# Sort by Amount_sum in ascending order
grouped_data = grouped_data.sort_values(by='Total Sales Value', ascending=False)

# Add Sr. No. column
grouped_data.insert(0, 'SR. No.', range(1, len(grouped_data) + 1))

# ✅ Customer-wise summary from grouped_data
customer_item_pairs = []

for _, row in grouped_data.iterrows():
    item_name = row['Item Name']
    customers = row['List of Customers'].split(' / ')
    total_value = row['Total Sales Value']
    # ✅ Divide equally among customers
    split_value = total_value / len(customers) if len(customers) > 0 else 0

    for cust in customers:
        customer_item_pairs.append([cust.strip(), item_name, split_value])

# ✅ Create DataFrame
customer_item_df = pd.DataFrame(customer_item_pairs, columns=['Customer', 'Item Name', 'Value'])

# ✅ Group by Customer
customer_summary = customer_item_df.groupby('Customer').agg(
    Unique_Items=('Item Name', pd.Series.nunique),
    Items_List=('Item Name', lambda x: ' | '.join(sorted(x.unique()))),
    Total_Sales_Value=('Value', 'sum')
).reset_index()

# ✅ Add Sr. No.
customer_summary.insert(0, 'SR. No.', range(1, len(customer_summary) + 1))

print("✅ Customer-wise summary:")
print(customer_summary)

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/03_List of items where Dependency on 2-3 customer only.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    grouped_data.to_excel(writer, sheet_name='2 & 3 Customers', index=False)
    customer_summary.to_excel(writer, sheet_name='Customer Wise Summary', index=False)

print("✅ Data successfully saved to new file:", output_path)

