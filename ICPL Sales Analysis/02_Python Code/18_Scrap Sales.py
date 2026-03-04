import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime

# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# ✅ Filter data for "Voucher Type" == "Scrap Sales 2024-25"
filtered_data = data[data['Voucher Type'] == 'Scrap Sales 2025-26']

# ✅ Convert 'Rate' to numeric
filtered_data['Rate'] = pd.to_numeric(filtered_data['Rate'], errors='coerce')

# ✅ Group by 'Item Name'
summary = filtered_data.groupby('Item Name').agg(
    Unique_Rate_Count=('Rate', lambda x: x.nunique()),
   Unique_Rate_List=('Rate', lambda x: '|'.join(str(rate) for rate in sorted(x.dropna().unique()))),
    Max_Rate=('Rate', 'max'),
    Min_Rate=('Rate', 'min'),
    Mean_Rate=('Rate', 'mean')
).reset_index()

# ✅ Calculate formula column
summary['Difference Percentage'] = ((summary['Max_Rate'] - summary['Mean_Rate']) / summary['Max_Rate']) * 100

summary = summary[summary['Difference Percentage'] >= 10]

# ✅ Round values for neatness
summary['Mean_Rate'] = summary['Mean_Rate'].round(2)
summary['Difference Percentage'] = summary['Difference Percentage'].round(2)

# ✅ Add SR No. column starting from 1
summary.insert(0, 'SR No.', range(1, len(summary) + 1))

# ✅ Rename columns as per your requirement
summary.rename(columns={
    'Unique_Rate_Count': 'Rate Value Count',
    'Unique_Rate_List': 'List of Different Rate Value',
    'Max_Rate': 'Max Rate Value',
    'Min_Rate': 'Min Rate Value',
    'Mean_Rate': 'Mean(Rate)'
}, inplace=True)

# ✅ Print the final summary
print(summary)

# ✅ Get item names from summary
item_names = summary['Item Name'].tolist()

# ✅ Get all records for those item names from original filtered data
detailed_records = filtered_data[filtered_data['Item Name'].isin(item_names)]

# ✅ Add SR No. column starting from 1
detailed_records.insert(0, 'SR No.', range(1, len(detailed_records) + 1))

print(detailed_records)

# ✅ Optional: Export to Excel
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/18_Scrap Sales.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    summary.to_excel(writer, index=False, sheet_name='Rate Summary', startrow=2)
    detailed_records.to_excel(writer, index=False, sheet_name='Detailed Records', startrow=2)