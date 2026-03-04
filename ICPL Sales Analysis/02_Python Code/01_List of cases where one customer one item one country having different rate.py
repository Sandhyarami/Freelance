import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime

# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# Clean and standardize relevant columns
data['Customer'] = data['Customer'].astype(str).str.strip()
data['Item Name'] = data['Item Name'].astype(str).str.strip()
data['Country To'] = data['Country To'].astype(str).str.strip()

# ✅ Replace blank or NaN with 'BLANK' so they are included in groupby
data['Item Name'] = data['Item Name'].replace('', 'BLANK').fillna('BLANK')
data['Country To'] = data['Country To'].replace('', 'BLANK').fillna('BLANK')

# Ensure relevant columns exist and clean them
data['Customer'] = data['Customer'].astype(str).str.strip()
data['Item Name'] = data['Item Name'].astype(str).str.strip()
data['Rate'] = pd.to_numeric(data['Rate'], errors='coerce')
data['Quantity'] = pd.to_numeric(data['Quantity'], errors='coerce')
data['Date'] = pd.to_datetime(data['Date'], errors='coerce')

# ✅ Remove Quantity = 0
data = data[data['Quantity'] > 0]
    
# Function to return "value (count)" joined by |
def list_with_count(series):
    return ' | '.join([f"{x} ({c})" for x, c in series.value_counts().items()])

# Group by Customer and Item Name
grouped = data.groupby(['Customer', 'Item Name','Country To'])

# Aggregation
result = grouped.agg(
    Rate_Unique_Count=('Rate', lambda x: x.nunique()),
    Rate_List_Count=('Rate', list_with_count),
    Quantity_Unique_Count=('Quantity', lambda x: x.nunique()),
    Quantity_List_Count=('Quantity', list_with_count),
    Date_Unique_Count=('Date', lambda x: x.dt.date.nunique()),
    Quantity_List_By_Date=('Quantity', list_with_count),  # If needed separately for date, clarify logic
    Rate_Max=('Rate', 'max'),
    Rate_Min=('Rate', 'min')
).reset_index()

# Add Rate Difference column
result['Rate_Max_Min_Diff'] = result['Rate_Max'] - result['Rate_Min']

# ✅ Filter where unique rate count >= 2
filtered_result = result[result['Rate_Unique_Count'] >= 2]

# ✅ Add SR No. column starting from 1
filtered_result.insert(0, 'SR No.', range(1, len(filtered_result) + 1))

# ✅ Rename columns as per your requirement
filtered_result.rename(columns={
    'Customer': 'Customer Name',
    'Item Name': 'Item Name',
    'Country To': 'Country To',
    'Rate_Unique_Count': 'Count of Rate',
    'Rate_List_Count': 'List of Rate with Count',
    'Quantity_Unique_Count': 'Count of Quantity',
    'Quantity_List_Count': 'List of Quantity with Count',
    'Date_Unique_Count': 'Count of Sales Date',
    'Quantity_List_By_Date': 'List of Sales Date with Count',
    'Rate_Max': 'Max*(Rate)',
    'Rate_Min': 'Min*(Rate)',
    'Rate_Max_Min_Diff': 'Max-Min'
}, inplace=True)

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/01_List of cases where one customer one item one country having different rate.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    filtered_result.to_excel(writer, sheet_name='Multiple Rate', index=False,startrow=2)

print("✅ Data successfully saved to new file:", output_path)
