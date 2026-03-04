import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
import re

# ✅ Load Excel workbook
file_path = load_workbook(
    r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx")

print("Sheet names:", file_path.sheetnames)

sheet = file_path['Working']

# ✅ Load sheet into DataFrame
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')
print("✅ Raw Data Preview:")
print(data.head())

# ✅ Convert Date columns to datetime
data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
data['Delivery Date'] = pd.to_datetime(data['Delivery Date'], errors='coerce')

# ✅ Calculate difference in days
data['Date_Diff_Days'] = (data['Delivery Date'] - data['Date']).dt.days

# ✅ Dynamic bucket function with NO limit
def get_bucket(days, step=30):
    if pd.isna(days):
        return 'NA'
    elif days == 0:
        return '0'
    elif days > 0:
        bucket_start = (days // step) * step + 1
        bucket_end = bucket_start + step - 1
        if bucket_start == 1:
            bucket_start = 0  # Pehla bucket 0 se shuru
        return f'{bucket_start}-{bucket_end}'
    else:
        # For negative days, bucket in reverse
        abs_days = abs(days)
        bucket_start = -((abs_days // step) * step + 1)
        bucket_end = bucket_start + step - 1
        return f'{bucket_start}-{bucket_end}'

# ✅ Apply bucket
data['Date_Diff_Bucket'] = data['Date_Diff_Days'].apply(lambda x: get_bucket(x, step=30))

print("\n✅ Preview with Difference & Bucket:")
print(data[['Date', 'Delivery Date', 'Date_Diff_Days', 'Date_Diff_Bucket']].head())

# ✅ Optional: Save to new Excel
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/25_Gap between Billing date and delivery date.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    data.to_excel(writer, sheet_name='Date Diff & Bucket', index=False)

print(f"\n✅ File saved at: {output_path}")
