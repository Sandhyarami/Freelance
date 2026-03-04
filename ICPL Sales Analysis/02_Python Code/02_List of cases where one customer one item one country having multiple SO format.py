import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
import re
# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

# ✅ Remove blank Order No rows
data = data[~data['Order No.'].isna() & (data['Order No.'].astype(str).str.strip() != '')]

print(data)

# ✅ Add "IN Format" column with classification
def classify_order_no(val):
    val = str(val).strip()
    if re.match(r'^[a-zA-Z0-9/-]+$', val) and not re.match(r'^[0-9]+$', val):
        return "Alphanumeric"
    elif re.match(r'^[0-9]+$', val):
        return "Numeric"
    else:
        return "Other"

data['IN Format'] = data['Order No.'].apply(classify_order_no)

# ✅ Group and create summary
def format_with_count(series):
    return ' | '.join([f"{k} ({v})" for k, v in series.value_counts().items()])

grouped = data.groupby(['Customer', 'Item Name', 'Country To']).agg(
    IN_Format_Unique_Count=('IN Format', 'nunique'),
    IN_Format_List_with_Count=('IN Format', format_with_count),
    Order_No_List=('Order No.', lambda x: ' | '.join(sorted(x.astype(str).unique())))
).reset_index()

grouped = grouped[grouped['IN_Format_Unique_Count'] >= 2]

# ✅ Rename columns as per your requirement
grouped.rename(columns={
    'IN_Format_Unique_Count': 'Count of Format',
    'IN_Format_List_with_Count': 'Type of Format',
    'Order_No_List' : 'List of Sales Order No.'
}, inplace=True)

# 🔢 Add SR No. starting from 1
grouped.insert(0, 'SR No.', range(1, len(grouped) + 1))

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/02_List of cases where one customer one item one country having multiple SO format.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    grouped.to_excel(writer, sheet_name='Multiple IN Format', index=False)

print("✅ Data successfully saved to new file:", output_path)