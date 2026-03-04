import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Folder path where Excel files are stored
folder_path = r"D:/Sandhya/Analysis Insights/Data analysis/DKFPL/Q1 2025/DKFPL Q1 Insights/02_Output Data - Copy"

# Define thin black border
thin_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# Define fills
title_fill = PatternFill(start_color='E6B8B7', end_color='E6B8B7', fill_type='solid')  # RGB(230,184,183)
header_fill = PatternFill(start_color='C4BD97', end_color='C4BD97', fill_type='solid')  # RGB(196,189,151)

for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
        file_path = os.path.join(folder_path, file_name)
        print(f"Processing: {file_name}")

        wb = load_workbook(file_path)
        for sheet in wb.worksheets:
            max_col = sheet.max_column
            max_row = sheet.max_row
            
            # 🚫 Hide gridlines
            sheet.sheet_view.showGridLines = False

            # 🔁 Step 1: Safely unmerge existing merged cells
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                try:
                    sheet.unmerge_cells(merged_range.coord)
                except:
                    print(f"⚠️ Could not unmerge: {merged_range.coord}")

            # 🔹 Step 2: Merge row 7-8 (A7 to last column)
            merged_value = sheet.cell(row=7, column=1).value
            sheet.merge_cells(start_row=7, start_column=1, end_row=8, end_column=max_col)
            cell = sheet.cell(row=7, column=1)
            cell.value = merged_value

            # ✅ Title styling
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=14)
            cell.fill = title_fill

            # ✅ Border for merged cells (row 7 & 8)
            for row in [7, 8]:
                for col in range(1, max_col + 1):
                    sheet.cell(row=row, column=col).border = thin_border

            # ✅ Header row (row 9)
            for col in range(1, max_col + 1):
                header_cell = sheet.cell(row=9, column=col)
                header_cell.font = Font(bold=True, size=11)
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.fill = header_fill
                header_cell.border = thin_border

            # ✅ Data cells from row 10 onward
            for row in sheet.iter_rows(min_row=10, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = thin_border

            # ✅ Auto-fit columns
            for col in range(1, max_col + 1):
                max_length = 0
                col_letter = sheet.cell(row=9, column=col).column_letter
                for row in range(7, max_row + 1):
                    try:
                        cell = sheet.cell(row=row, column=col)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_path)

print("✅ Done: All formatting applied after clearing previous merges.")
