import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font , PatternFill
from datetime import datetime
from openpyxl.utils import get_column_letter
import re
# Example usage
file_path = load_workbook(r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/01_Input Data/02_Clean Data/Sales Data Clean.xlsx") # Replace with the path to your file

# Print the sheet names
print("Sheet names:", file_path.sheetnames)

# Access a specific sheet by name
sheet = file_path['Working']

# Load data into a DataFrame using pandas
data = pd.read_excel(file_path, sheet_name='Working', engine='openpyxl')

print(data)

# ✅ Group and create summary
def format_with_count(series):
    return ' | '.join([f"{k} ({v})" for k, v in series.value_counts().items()])

grouped = data.groupby(['Order No.']).agg(
    Order_Date_Unique_Count=('Order Date', 'nunique'),
    Order_Date_List_with_Count=('Order Date', format_with_count),
    
).reset_index()

grouped = grouped[grouped['Order_Date_Unique_Count'] >= 2]

# ✅ Rename columns as per your requirement
grouped.rename(columns={
    'Order No.' : 'Sales Order No.',
    'Order_Date_Unique_Count': 'Count of Order Date',
    'Order_Date_List_with_Count': 'List of Order Date with Count'
}, inplace=True)

# 🔢 Add SR No. starting from 1
grouped.insert(0, 'SR No.', range(1, len(grouped) + 1))

# ✅ Save to a new Excel file
output_path = r"D:/Sandhya/Analysis Insights/Data analysis/ICPL/31-7-2025/02_Output Data/05_List of sales order number having multiple Order Date.xlsx"
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    grouped.to_excel(writer, sheet_name='Multiple Order Date', index=False)

print("✅ Data successfully saved to new file:", output_path)